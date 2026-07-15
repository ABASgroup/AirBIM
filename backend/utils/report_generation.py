from uuid import UUID
import json
import re
from pathlib import Path
from datetime import datetime
from collections.abc import Mapping
from functools import lru_cache
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors


_RU_REPORT_TYPES = {
    "progress": "Отчет о прогрессе",
    "plan_fact": "Отчет по план-факту",
}

_RU_LABELS = {
    "parameter": "Параметр",
    "value": "Значение",
    "project": "Проект",
    "project_name": "Название проекта",
    "project_description": "Описание проекта",
    "project_id": "ID проекта",
    "stage_id": "ID этапа",
    "old_stage_id": "ID старого этапа",
    "new_stage_id": "ID нового этапа",
    "stage": "Этап",
    "stage_name": "Название этапа",
    "stage_description": "Описание этапа",
    "old_stage_name": "Название старого этапа",
    "old_stage_description": "Описание старого этапа",
    "new_stage_name": "Название нового этапа",
    "new_stage_description": "Описание нового этапа",
    "stage_start_date": "Дата начала этапа",
    "old_stage_start_date": "Дата начала старого этапа",
    "new_stage_start_date": "Дата начала нового этапа",
    "name": "Название",
    "description": "Описание",
    "age": "Возраст",
    "status": "Статус",
    "type": "Тип",
    "images": "Изображения",
    "report": "Отчет",
    "data": "Данные",
    "tolerance": "Допуск",
}


def translate_recording_result_type(result_type: str) -> str:
    """Translate recording result type into a Russian report title."""
    normalized = str(result_type).strip().lower()
    return _RU_REPORT_TYPES.get(normalized, normalized.replace("_", " ").capitalize())


def _value_cleaner(value):
    """
    Cleans data before you need to pass it into a report.

    Returns:
        Unknown | str: your value if no changes were applied or your value in str format
    """
    if isinstance(value, (list, tuple, set)):
        return ", ".join(map(str, value))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    return value


def _label_cleaner(value: str) -> str:
    """Translate common labels to Russian and keep unknown ones readable."""
    normalized = str(value).strip().lower()
    if normalized in _RU_LABELS:
        return _RU_LABELS[normalized]
    key_match = re.fullmatch(r"key(\d+)", normalized)
    if key_match:
        return f"Ключ {key_match.group(1)}"
    return str(value).replace("_", " ").capitalize()


def _prepare_pairs(data: Mapping[str, object] | None) -> list[tuple[str, object]]:
    if not data:
        return []
    return [(str(key), value) for key, value in data.items()]


def extract_report_sections(
    data: Mapping[str, object],
    section_specs: dict[str, list[tuple[str, str]]],
) -> tuple[dict[str, dict[str, object]], dict[str, object]]:
    """Split report metadata sections from the main data dictionary."""
    remaining = dict(data)
    sections: dict[str, dict[str, object]] = {}

    for section_title, entries in section_specs.items():
        section_rows: dict[str, object] = {}
        for source_key, target_label in entries:
            if source_key in remaining:
                section_rows[target_label] = remaining.pop(source_key)

        if section_rows:
            sections[section_title] = section_rows

    return sections, remaining


def _coerce_path(file_path: Path | str) -> Path:
    return file_path if isinstance(file_path, Path) else Path(file_path)


@lru_cache(maxsize=1)
def _register_pdf_fonts() -> tuple[str, str]:
    candidates = [
        (
            "Arial",
            Path(r"C:\Windows\Fonts\arial.ttf"),
            Path(r"C:\Windows\Fonts\arialbd.ttf"),
        ),
        (
            "DejaVuSans",
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ),
        (
            "LiberationSans",
            Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
            Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"),
        ),
    ]

    for family_name, regular_path, bold_path in candidates:
        if regular_path.exists() and bold_path.exists():
            regular_font_name = f"{family_name}Regular"
            bold_font_name = f"{family_name}Bold"
            pdfmetrics.registerFont(TTFont(regular_font_name, str(regular_path)))
            pdfmetrics.registerFont(TTFont(bold_font_name, str(bold_path)))
            return regular_font_name, bold_font_name

    return "Helvetica", "Helvetica-Bold"


def _build_pdf_table(rows: list[tuple[str, object]], styles, header_font_name: str) -> Table:
    table_data = [[_label_cleaner("parameter"), _label_cleaner("value")]]
    table_data.extend([
        [Paragraph(_label_cleaner(key), styles["BodyText"]), Paragraph(str(_value_cleaner(value)), styles["BodyText"])]
        for key, value in rows
    ])

    table = Table(table_data, colWidths=[175, 330], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), header_font_name),
        ('FONTSIZE',   (0, 0), (-1, 0), 11),

        ('FONTSIZE',   (0, 1), (-1, -1), 10),
        ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
        ('ALIGN',      (0, 0), (-1, -1), 'LEFT'),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor("#C7D3E0")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")]),

        ('TOPPADDING',    (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
    ]))
    return table


def _build_pdf_story_block(story, section_title: str, rows: list[tuple[str, object]], styles):
    if not rows:
        return

    story.append(Paragraph(section_title, styles['ReportSection']))
    story.append(Spacer(1, 6))
    story.append(_build_pdf_table(rows, styles, styles['ReportSection'].fontName))
    story.append(Spacer(1, 12))


def _style_excel_header(cell):
    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_excel_section_title(cell):
    cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_excel_body(cell, *, bold: bool = False, fill: str = "FFFFFF"):
    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )


def _write_excel_kv_section(sheet, start_row: int, title: str, rows: list[tuple[str, object]]) -> int:
    if not rows:
        return start_row

    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    title_cell = sheet.cell(row=start_row, column=1, value=title)
    _style_excel_section_title(title_cell)

    current_row = start_row + 1
    for key, value in rows:
        label_cell = sheet.cell(row=current_row, column=1, value=_label_cleaner(key))
        value_cell = sheet.cell(row=current_row, column=2, value=_value_cleaner(value))
        _style_excel_body(label_cell, bold=True, fill="F4F8FC")
        _style_excel_body(value_cell)
        current_row += 1

    return current_row + 1


def _write_excel_horizontal_table(sheet, start_row: int, rows: list[tuple[str, object]]) -> int:
    if not rows:
        return start_row

    for column, (header, value) in enumerate(rows, start=1):
        header_cell = sheet.cell(row=start_row, column=column, value=_label_cleaner(header))
        value_cell = sheet.cell(row=start_row + 1, column=column, value=_value_cleaner(value))
        _style_excel_header(header_cell)
        _style_excel_body(value_cell)

    return start_row + 2


def generate_excel_report(
    title: str,
    data: dict,
    file_path: Path | str,
    sections: dict[str, dict[str, object]] | None = None,
):
    """
    Generates an Excel (`.xlsx`) report with the given title and data.

    Saves your data in two rows (headers and values) as a table.

    The report will be saved at the specified file path (use `.xlsx` format).

    Args:
        title (str): the title that will be used for the report
        data (dict): dict data that will be placed in the table
        file_path (Path | str): where you need to save the path
    """

    file_path = _coerce_path(file_path)

    workbook = Workbook()
    sheet = workbook.active

    # title
    sheet.title = title

    data_rows = _prepare_pairs(data)
    section_rows = [(section_title, _prepare_pairs(section_data)) for section_title, section_data in (sections or {}).items()]
    max_columns = max(2, len(data_rows))

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_columns)
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 3

    for section_title, rows in section_rows:
        current_row = _write_excel_kv_section(sheet, current_row, section_title, rows)

    _write_excel_horizontal_table(sheet, current_row, data_rows)

    # adjust column widths
    for column_index in range(1, max_columns + 1):
        col_letter = get_column_letter(column_index)
        max_len = 0
        for row in sheet.iter_rows(min_col=column_index, max_col=column_index):
            for cell in row:
                max_len = max(max_len, len(str(cell.value or "")))
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # save on the path
    if file_path.exists():
        # remove existing file to avoid overwrite issues
        # maybe need to create copy of the file instead of deleting it?
        file_path.unlink()
    workbook.save(file_path)


def generate_pdf_report(
    title: str,
    data: dict,
    file_path: Path | str,
    imgs: list[Path] | list[str] | None = None,
    sections: dict[str, dict[str, object]] | None = None,
):
    """
    Generates a PDF report with the given title and data.

    The report will be saved at the specified file path (use `.pdf` format).

    Args:
        title (str): the title that will be used for the report
        data (dict): dict data that will be placed in the table
        file_path (Path | str): where you need to save the path
        imgs (list[Path] | list[str] | None, optional): the images you want to add to the report. Defaults to None.
    """
    file_path = _coerce_path(file_path)
    pdf_body_font_name, pdf_bold_font_name = _register_pdf_fonts()

    document = SimpleDocTemplate(
        filename=str(file_path),
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    # styles object
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="ReportTitle",
        parent=styles["Title"],
        fontName=pdf_bold_font_name,
        fontSize=20,
        leading=24,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=10,
    ))
    styles.add(ParagraphStyle(
        name="ReportSection",
        parent=styles["Heading2"],
        fontName=pdf_bold_font_name,
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#1F4E79"),
        spaceBefore=10,
        spaceAfter=6,
    ))
    styles["BodyText"].fontName = pdf_body_font_name
    styles["BodyText"].fontSize = 10
    styles["BodyText"].leading = 12

    # document's elements
    story = []

    # add title
    story.append(Paragraph(title, styles['ReportTitle']))
    story.append(Spacer(1, 10))

    for section_title, section_data in (sections or {}).items():
        _build_pdf_story_block(story, section_title, _prepare_pairs(section_data), styles)

    data_rows = _prepare_pairs(data)
    if data_rows:
        story.append(Paragraph(_label_cleaner("data"), styles['ReportSection']))
        story.append(_build_pdf_table(data_rows, styles, pdf_bold_font_name))
        story.append(Spacer(1, 12))

    # add optional images
    if imgs is not None and len(imgs) != 0:
        story.append(PageBreak())
        story.append(Paragraph(f"{title}: {_label_cleaner('images').lower()}", styles['ReportTitle']))
        for img in imgs:
            img = Image(str(img))
            # scale
            width = document.width * 0.7
            scale = width / img.imageWidth
            img.drawWidth = width
            img.drawHeight = img.imageHeight * scale

            img.hAlign = 'CENTER'
            story.append(img)
            story.append(Spacer(1, 10))

    # save on the path
    document.build(story)
