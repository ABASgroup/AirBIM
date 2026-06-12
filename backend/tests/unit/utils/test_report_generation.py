"""Unit tests for utils.report_generation."""

from pathlib import Path
from typing import Any

import openpyxl
import pytest
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pytest_mock import MockerFixture
from reportlab.lib import colors
from reportlab.platypus import PageBreak, SimpleDocTemplate, Table

from utils.report_generation import generate_excel_report, generate_pdf_report


def test_basic_report(tmp_path: Path) -> None:
    """Correct structure: headers in row 1, values in row 2."""
    file_path = tmp_path / "report.xlsx"
    data = {"Name": "Alice", "Age": 30}
    generate_excel_report("Test", data, file_path)

    assert file_path.exists()
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    assert ws.title == "Test"
    assert ws.cell(1, 1).value == "Name"
    assert ws.cell(2, 1).value == "Alice"
    assert ws.cell(1, 2).value == "Age"
    assert ws.cell(2, 2).value == 30


def test_empty_data_creates_empty_sheet(tmp_path: Path) -> None:
    """Empty dict creates a file with sheet title but no data rows."""
    file_path = tmp_path / "empty.xlsx"
    generate_excel_report("Empty", {}, file_path)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    assert ws.title == "Empty"
    # No data rows — max_row may be None or 1 depending on openpyxl version
    assert ws.max_row is None or ws.max_row == 1


def test_style_and_width(tmp_path: Path) -> None:
    """Headers get yellow fill, column widths are calculated correctly."""
    file_path = tmp_path / "styled.xlsx"
    # "Short" (len 5) vs "1" (len 1) -> max 5, width = 5+3=8 -> clamped to min 11
    # "LongerHeader" (len 12) vs "12345" (len 5) -> max 12, width = 15
    data = {"Short": 1, "LongerHeader": 12345}
    generate_excel_report("Style", data, file_path)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    expected_fill = PatternFill(
        start_color="FFCF40", end_color="FFCF40", fill_type="solid"
    )
    for col in range(1, 3):
        cell = ws.cell(1, col)
        assert cell.fill.start_color.rgb == expected_fill.start_color.rgb
        assert cell.fill.end_color.rgb == expected_fill.end_color.rgb

    assert ws.column_dimensions[get_column_letter(1)].width == 11
    assert ws.column_dimensions[get_column_letter(2)].width == 15


def test_basic_pdf_creation(tmp_path: Path) -> None:
    """A valid PDF file is created and is non-empty."""
    file_path = tmp_path / "report.pdf"
    data = {"Name": "Alice", "Age": 30}
    generate_pdf_report("Test Report", data, file_path)

    assert file_path.exists()
    assert file_path.stat().st_size > 0


def test_empty_data_creates_pdf(tmp_path: Path) -> None:
    """Empty data dict still produces a valid PDF file."""
    file_path = tmp_path / "empty.pdf"
    generate_pdf_report("Empty", {}, file_path)

    assert file_path.exists()
    assert file_path.stat().st_size > 0


def test_value_cleaner_called_for_all_values(
    tmp_path: Path, mocker: MockerFixture
) -> None:
    """_value_cleaner is called for each value in the data dict."""
    mock_cleaner = mocker.patch(
        "utils.report_generation._value_cleaner", side_effect=lambda x: f"clean_{x}"
    )

    data = {"A": 1, "B": [2, 3], "C": None}
    file_path = tmp_path / "cleaner.pdf"
    generate_pdf_report("Test", data, file_path)

    assert mock_cleaner.call_count == len(data)
    mock_cleaner.assert_any_call(1)
    mock_cleaner.assert_any_call([2, 3])
    mock_cleaner.assert_any_call(None)


def test_table_creation_and_styling(tmp_path: Path, mocker: MockerFixture) -> None:
    """Verify that Table is created with correct data and style."""
    mock_table = mocker.patch("utils.report_generation.Table", wraps=Table)
    mock_style = mocker.patch("utils.report_generation.TableStyle")

    data = {"Key1": "Val1", "Key2": "Val2"}
    file_path = tmp_path / "table.pdf"
    generate_pdf_report("Table Test", data, file_path)

    # Check that Table was called with headers + data rows
    expected_data = [
        ["Parameter", "Value"],
        ["Key1", "Val1"],
        ["Key2", "Val2"],
    ]
    args, _ = mock_table.call_args
    assert args[0] == expected_data

    # Check that TableStyle was called with styling commands
    style_call_args = mock_style.call_args[0][0]
    assert any(
        cmd[0] == "BACKGROUND" and cmd[3] == colors.HexColor("#2C3E50")
        for cmd in style_call_args
    )
    assert any(
        cmd[0] == "TEXTCOLOR" and cmd[3] == colors.white for cmd in style_call_args
    )


def test_images_added_when_provided(tmp_path: Path, mocker: MockerFixture) -> None:
    """When images list is not empty, PageBreak and Images are added to the story."""
    mock_image = mocker.patch("utils.report_generation.Image")
    # Mock Image instance attributes to avoid file access
    mock_img_instance = mock_image.return_value
    mock_img_instance.imageWidth = 100
    mock_img_instance.imageHeight = 200

    # Intercept document.build to inspect the story
    original_build = SimpleDocTemplate.build
    story_elements = []

    def fake_build(
        self: SimpleDocTemplate,
        story: list[Any],
        **kwargs: Any,
    ) -> Any:
        story_elements.extend(story)
        # Still write the file for the assertion
        return original_build(self, story, **kwargs)

    mocker.patch.object(SimpleDocTemplate, "build", fake_build)

    file_path = tmp_path / "with_images.pdf"
    imgs = ["img1.png", "img2.jpg"]
    generate_pdf_report("Image Test", {"A": 1}, file_path, imgs=imgs)

    # Verify PageBreak was added
    assert any(isinstance(el, PageBreak) for el in story_elements)
    # Verify Image was called for each path
    assert mock_image.call_count == 2
    mock_image.assert_any_call("img1.png")
    mock_image.assert_any_call("img2.jpg")

    # Verify images are in the story
    image_elements = [el for el in story_elements if el is mock_img_instance]
    assert len(image_elements) == 2


def test_missing_directory_raises(tmp_path: Path) -> None:
    """A FileNotFoundError is raised if the target directory does not exist."""
    bad_path = tmp_path / "missing" / "report.pdf"
    with pytest.raises(FileNotFoundError):
        generate_pdf_report("Bad", {"A": 1}, bad_path)
